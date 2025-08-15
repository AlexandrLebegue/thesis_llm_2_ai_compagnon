"""
Excel Generator Tool for Smol Agents

This tool allows the agent to create Excel files with data, tables, and charts
based on user requests and data analysis.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import tempfile
import os
import json
import logging
from typing import Dict, List, Any, Optional
from smolagents import Tool

# Import our robust utility functions
from .tool_utils import (
    ToolInputSanitizer, ToolValidator, FileVerifier,
    DebugLogger, ErrorFormatter
)
from .excel_preview import ExcelPreviewGenerator

logger = logging.getLogger(__name__)


class ExcelGeneratorTool(Tool):
    """
    A tool that allows the agent to create Excel files with structured data, formatting, and charts.
    
    This tool can:
    - Create Excel workbooks with multiple sheets
    - Add formatted tables and data
    - Generate charts and visualizations
    - Apply professional styling
    """
    
    name = "excel_generator"
    description = """
    Generate Excel files (.xlsx) with structured data, tables, and charts.
    
    This tool can create professional Excel documents with:
    - Multiple worksheets
    - Formatted data tables
    - Charts and visualizations (bar, line, pie charts)
    - Professional styling and formatting
    - Headers, footers, and metadata
    
    Use this tool when you need to create structured data reports, analysis results,
    or any tabular data that would benefit from Excel formatting.
    
    Example usage:
    excel_generator(
        data_structure='{"sheets": [{"name": "Sales", "tables": [{"data": [["Product", "Sales"], ["Widget A", 100], ["Widget B", 150]]}]}]}',
        filename="sales_report"
    )
    
    The data_structure should be a JSON string. Both single and double quotes are supported.
    """
    
    inputs = {
        "data_structure": {
            "type": "string",
            "description": "JSON string describing the Excel structure with sheets, data, and formatting"
        },
        "filename": {
            "type": "string",
            "description": "Name for the Excel file (without extension)"
        }
    }
    
    output_type = "string"
    
    def forward(self, data_structure: str, filename: str) -> str:
        """
        Generate an Excel file based on the provided data structure.
        
        Args:
            data_structure (str): JSON string describing the Excel structure
            filename (str): Name for the Excel file
            
        Returns:
            str: Path to the generated Excel file or error message
        """
        # Initialize debug logging
        DebugLogger.log_tool_start('excel_generator', {
            'data_structure_length': len(str(data_structure)),
            'filename': filename
        })
        
        try:
            # Robust JSON parsing with multiple fallback strategies
            structure = ToolInputSanitizer.safe_json_loads(
                data_structure,
                fallback_value={'sheets': []}
            )
            
            DebugLogger.log_json_parsing('excel_generator', str(data_structure), structure)
            
            # Validate input structure
            is_valid, validation_error = ToolValidator.validate_input(structure, 'excel_generator')
            if not is_valid:
                DebugLogger.log_validation_result('excel_generator', False, validation_error)
                return ErrorFormatter.format_validation_error('excel_generator', validation_error)
            
            DebugLogger.log_validation_result('excel_generator', True)
            
            # Create Excel workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Process each sheet in the structure
            sheets_created = 0
            for sheet_config in structure.get('sheets', []):
                try:
                    self._create_sheet(workbook, sheet_config)
                    sheets_created += 1
                    logger.debug(f"Created sheet: {sheet_config.get('name', 'Unnamed')}")
                except Exception as e:
                    logger.warning(f"Failed to create sheet {sheet_config.get('name', 'Unnamed')}: {str(e)}")
                    # Continue with other sheets
            
            # If no sheets were created, create a default one
            if not workbook.worksheets:
                logger.info("No sheets created from input, creating default sheet")
                self._create_default_sheet(workbook)
                sheets_created = 1
            
            # Save the workbook
            output_path = self._save_workbook(workbook, filename)
            
            # Verify file creation
            verification = FileVerifier.verify_excel_file(output_path)
            DebugLogger.log_file_operation(
                'excel_generator',
                'creation',
                output_path,
                verification['exists'] and verification['has_data'],
                f"Sheets: {sheets_created}, Size: {verification.get('size', 0)} bytes"
            )
            
            if not verification['exists']:
                return ErrorFormatter.format_file_creation_error('excel_generator', output_path, verification)
            
            if not verification['has_data'] and sheets_created > 0:
                logger.warning("Excel file created but appears to have no data")
                return f"Excel file created but may be empty: {output_path}"
            
            # Generate HTML preview
            preview_result = ExcelPreviewGenerator.generate_preview(output_path)
            preview_html = None
            if preview_result['success']:
                preview_html = preview_result['preview_html']
                logger.info(f"Generated HTML preview for {filename}")
            else:
                logger.warning(f"Failed to generate preview for {filename}: {preview_result.get('error', 'Unknown error')}")
            
            # Return structured result including preview
            result = {
                'file_path': output_path,
                'preview_html': preview_html,
                'message': f"Excel file created successfully: {output_path}"
            }
            
            return str(result)  # Convert to string for tool compatibility
            
        except Exception as e:
            logger.error(f"Error in excel_generator: {str(e)}", exc_info=True)
            if "json" in str(e).lower():
                return ErrorFormatter.format_json_error('excel_generator', str(data_structure), e)
            return f"Error creating Excel file: {str(e)}"
    
    def _create_sheet(self, workbook: openpyxl.Workbook, sheet_config: Dict[str, Any]) -> None:
        """Create a worksheet based on the configuration."""
        sheet_name = sheet_config.get('name', 'Sheet1')
        sheet = workbook.create_sheet(title=sheet_name)
        
        # Add title if specified
        title = sheet_config.get('title')
        if title:
            sheet['A1'] = title
            sheet['A1'].font = Font(size=16, bold=True)
            sheet['A1'].alignment = Alignment(horizontal='center')
            sheet.merge_cells('A1:E1')
            current_row = 3
        else:
            current_row = 1
        
        # Add data tables
        for table_config in sheet_config.get('tables', []):
            current_row = self._add_table(sheet, table_config, current_row)
            current_row += 2  # Add spacing between tables
        
        # Add charts
        for chart_config in sheet_config.get('charts', []):
            self._add_chart(sheet, chart_config)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(sheet)
    
    def _add_table(self, sheet: openpyxl.worksheet.worksheet.Worksheet,
                   table_config: Dict[str, Any], start_row: int) -> int:
        """Add a data table to the worksheet with robust data handling."""
        data = table_config.get('data', [])
        headers = table_config.get('headers', [])
        table_title = table_config.get('title')
        
        current_row = start_row
        
        # Add table title
        if table_title:
            sheet.cell(row=current_row, column=1, value=table_title)
            sheet.cell(row=current_row, column=1).font = Font(size=12, bold=True)
            current_row += 2
        
        # Handle various data formats
        processed_data = self._normalize_table_data(data, headers)
        
        if not processed_data:
            logger.warning("No valid data found for table")
            return current_row
        
        # Extract headers and data rows
        if headers:
            table_headers = headers
            data_rows = processed_data
        elif processed_data and len(processed_data) > 0:
            # First row might be headers
            if isinstance(processed_data[0], (list, tuple)) and len(processed_data) > 1:
                # Check if first row looks like headers (contains strings)
                first_row = processed_data[0]
                if any(isinstance(cell, str) and not cell.isdigit() for cell in first_row):
                    table_headers = first_row
                    data_rows = processed_data[1:]
                else:
                    table_headers = []
                    data_rows = processed_data
            else:
                table_headers = []
                data_rows = processed_data
        else:
            table_headers = []
            data_rows = []
        
        # Add headers
        if table_headers:
            for col, header in enumerate(table_headers, 1):
                cell = sheet.cell(row=current_row, column=col, value=str(header))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            current_row += 1
        
        # Add data rows
        for row_data in data_rows:
            if not isinstance(row_data, (list, tuple)):
                # Handle single values or dictionaries
                if isinstance(row_data, dict):
                    row_data = list(row_data.values())
                else:
                    row_data = [row_data]
            
            for col, value in enumerate(row_data, 1):
                try:
                    # Safely convert value to appropriate type
                    cell_value = self._safe_cell_value(value)
                    cell = sheet.cell(row=current_row, column=col, value=cell_value)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    # Format numbers
                    if isinstance(cell_value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
                except Exception as e:
                    logger.warning(f"Error setting cell value at row {current_row}, col {col}: {e}")
                    sheet.cell(row=current_row, column=col, value=str(value) if value is not None else "")
            
            current_row += 1
        
        logger.debug(f"Added table with {len(data_rows)} data rows")
        return current_row
    
    def _normalize_table_data(self, data: Any, headers: List = None) -> List[List]:
        """Normalize various data formats into a consistent list of lists."""
        if not data:
            return []
        
        # Handle pandas DataFrame
        if hasattr(data, 'values') and hasattr(data, 'columns'):
            # It's a DataFrame
            result = data.values.tolist()
            if not headers and hasattr(data, 'columns'):
                # Prepend column names as first row
                result.insert(0, list(data.columns))
            return result
        
        # Handle list of dictionaries
        if isinstance(data, list) and data and isinstance(data[0], dict):
            keys = list(data[0].keys())
            result = []
            if not headers:
                result.append(keys)  # Add keys as headers
            for item in data:
                result.append([item.get(key, '') for key in keys])
            return result
        
        # Handle nested lists
        if isinstance(data, list):
            result = []
            for row in data:
                if isinstance(row, (list, tuple)):
                    result.append(list(row))
                elif isinstance(row, dict):
                    result.append(list(row.values()))
                else:
                    result.append([row])
            return result
        
        # Handle single values
        return [[data]]
    
    def _safe_cell_value(self, value: Any) -> Any:
        """Safely convert a value for Excel cell insertion."""
        if value is None:
            return ""
        
        # Handle numeric strings
        if isinstance(value, str):
            # Try to convert to number if it looks numeric
            try:
                if '.' in value:
                    return float(value)
                else:
                    return int(value)
            except ValueError:
                return value
        
        # Handle boolean values
        if isinstance(value, bool):
            return value
        
        # Handle numeric values
        if isinstance(value, (int, float)):
            return value
        
        # Convert everything else to string
        return str(value)
    
    def _add_chart(self, sheet: openpyxl.worksheet.worksheet.Worksheet, 
                   chart_config: Dict[str, Any]) -> None:
        """Add a chart to the worksheet."""
        chart_type = chart_config.get('type', 'bar')
        data_range = chart_config.get('data_range')
        title = chart_config.get('title', 'Chart')
        position = chart_config.get('position', 'G2')
        
        if not data_range:
            return
        
        # Create chart based on type
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            chart = BarChart()  # Default to bar chart
        
        # Set chart properties
        chart.title = title
        chart.style = 10
        
        # Add data to chart
        data = Reference(sheet, range_string=data_range)
        chart.add_data(data, titles_from_data=True)
        
        # Add chart to sheet
        sheet.add_chart(chart, position)
    
    def _auto_adjust_columns(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Auto-adjust column widths based on content."""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_default_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Create a default sheet with sample data."""
        sheet = workbook.create_sheet(title="Data")
        
        # Add sample headers
        headers = ["Item", "Value", "Category"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add sample data
        sample_data = [
            ["Sample Item 1", 100, "Category A"],
            ["Sample Item 2", 200, "Category B"],
            ["Sample Item 3", 150, "Category A"]
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        self._auto_adjust_columns(sheet)
    
    def _save_workbook(self, workbook: openpyxl.Workbook, filename: str) -> str:
        """Save the workbook and return the file path."""
        # Ensure filename has .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Create temp directory if it doesn't exist
        temp_dir = os.path.join(os.getcwd(), 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        
        # Save file
        output_path = os.path.join(temp_dir, filename)
        workbook.save(output_path)
        
        return output_path


class SimpleExcelGeneratorTool(Tool):
    """
    Simplified Excel generator for quick data export.
    """
    
    name = "simple_excel_generator"
    description = """
    Generate simple Excel files from tabular data.
    
    This tool creates basic Excel files from data arrays.
    Use this for quick data export without complex formatting.
    """
    
    inputs = {
        "data": {
            "type": "string",
            "description": "JSON string containing tabular data as array of arrays"
        },
        "headers": {
            "type": "string", 
            "description": "JSON string containing column headers"
        },
        "filename": {
            "type": "string",
            "description": "Name for the Excel file (without extension)"
        }
    }
    
    output_type = "string"
    
    def forward(self, data: str, headers: str, filename: str) -> str:
        """Generate a simple Excel file from data."""
        try:
            # Parse inputs
            data_array = json.loads(data)
            headers_array = json.loads(headers) if headers else None
            
            # Create DataFrame
            df = pd.DataFrame(data_array, columns=headers_array)
            
            # Ensure filename has .xlsx extension
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            # Create temp directory if it doesn't exist
            temp_dir = os.path.join(os.getcwd(), 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            
            # Save file
            output_path = os.path.join(temp_dir, filename)
            df.to_excel(output_path, index=False)
            
            return f"Simple Excel file created successfully: {output_path}"
            
        except Exception as e:
            return f"Error creating simple Excel file: {str(e)}"