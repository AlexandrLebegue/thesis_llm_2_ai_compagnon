"""
Excel Generator Tool for Smol Agents

This tool allows the agent to create Excel files with data, tables, and charts
based on user requests and data analysis.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import tempfile
import os
import json
from typing import Dict, List, Any, Optional
from smolagents import Tool


class ExcelGeneratorTool(Tool):
    """
    A tool that allows the agent to create Excel files with structured data, formatting, and charts.
    
    This tool can:
    - Create Excel workbooks with multiple sheets
    - Add formatted tables and data
    - Generate charts and visualizations
    - Apply professional styling
    """
    
    name = "excel_generator"
    description = """
    Create new Excel files (.xlsx) with multiple sheets, tables, and charts.
    
    Usage: excel_generator(data_structure="JSON_string", filename="report_name")
    
    Example - Simple Excel with one sheet:
    excel_generator(
        data_structure='{"sheets": [{"name": "Sales", "tables": [{"headers": ["Product", "Revenue"], "data": [["Widget A", 1000], ["Widget B", 1500]]}]}]}',
        filename="sales_report"
    )
    
    Example - Multi-sheet Excel with charts:
    excel_generator(
        data_structure='{"sheets": [{"name": "Data", "title": "Sales Report", "tables": [{"title": "Q1 Sales", "headers": ["Month", "Revenue"], "data": [["Jan", 1000], ["Feb", 1200], ["Mar", 1400]]}], "charts": [{"type": "bar", "title": "Monthly Revenue", "data_range": "A4:B7", "position": "E2"}]}, {"name": "Summary", "tables": [{"headers": ["Total", "Average"], "data": [[3600, 1200]]}]}]}',
        filename="quarterly_report"
    )
    
    JSON Format:
    {
        "sheets": [
            {
                "name": "Sheet1",
                "title": "Optional sheet title",
                "tables": [
                    {
                        "title": "Optional table title",
                        "headers": ["Col1", "Col2", "Col3"],
                        "data": [
                            ["Row1_Col1", "Row1_Col2", "Row1_Col3"],
                            ["Row2_Col1", "Row2_Col2", "Row2_Col3"]
                        ]
                    }
                ],
                "charts": [
                    {
                        "type": "bar|line|pie",
                        "title": "Chart Title",
                        "data_range": "A1:B10",
                        "position": "E2"
                    }
                ]
            }
        ]
    }
    
    Returns: "Excel file created successfully: temp/filename.xlsx"
    """
    
    inputs = {
        "data_structure": {
            "type": "string",
            "description": "JSON string with Excel structure (see description for format)"
        },
        "filename": {
            "type": "string",
            "description": "Name for the Excel file (without .xlsx extension)"
        }
    }
    
    output_type = "string"
    
    def forward(self, data_structure: str, filename: str) -> str:
        """
        Generate an Excel file based on the provided data structure.
        
        Args:
            data_structure (str): JSON string describing the Excel structure
            filename (str): Name for the Excel file
            
        Returns:
            str: Path to the generated Excel file
        """
        try:
            # Parse the data structure
            structure = json.loads(data_structure)
            
            # Create Excel workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Process each sheet in the structure
            for sheet_config in structure.get('sheets', []):
                self._create_sheet(workbook, sheet_config)
            
            # If no sheets were created, create a default one
            if not workbook.worksheets:
                self._create_default_sheet(workbook)
            
            # Save the workbook
            output_path = self._save_workbook(workbook, filename)
            
            return f"Excel file created successfully: {output_path}"
            
        except json.JSONDecodeError as e:
            return f"Error parsing data structure JSON: {str(e)}"
        except Exception as e:
            return f"Error creating Excel file: {str(e)}"
    
    def _create_sheet(self, workbook: openpyxl.Workbook, sheet_config: Dict[str, Any]) -> None:
        """Create a worksheet based on the configuration."""
        sheet_name = sheet_config.get('name', 'Sheet1')
        sheet = workbook.create_sheet(title=sheet_name)
        
        # Add title if specified
        title = sheet_config.get('title')
        if title:
            sheet['A1'] = title
            sheet['A1'].font = Font(size=16, bold=True)
            sheet['A1'].alignment = Alignment(horizontal='center')
            sheet.merge_cells('A1:E1')
            current_row = 3
        else:
            current_row = 1
        
        # Add data tables
        for table_config in sheet_config.get('tables', []):
            current_row = self._add_table(sheet, table_config, current_row)
            current_row += 2  # Add spacing between tables
        
        # Add charts
        for chart_config in sheet_config.get('charts', []):
            self._add_chart(sheet, chart_config)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(sheet)
    
    def _add_table(self, sheet: openpyxl.worksheet.worksheet.Worksheet, 
                   table_config: Dict[str, Any], start_row: int) -> int:
        """Add a data table to the worksheet."""
        data = table_config.get('data', [])
        headers = table_config.get('headers', [])
        table_title = table_config.get('title')
        
        current_row = start_row
        
        # Add table title
        if table_title:
            sheet.cell(row=current_row, column=1, value=table_title)
            sheet.cell(row=current_row, column=1).font = Font(size=12, bold=True)
            current_row += 2
        
        # Add headers
        if headers:
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            current_row += 1
        
        # Add data rows
        for row_data in data:
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                # Format numbers
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
            current_row += 1
        
        return current_row
    
    def _add_chart(self, sheet: openpyxl.worksheet.worksheet.Worksheet, 
                   chart_config: Dict[str, Any]) -> None:
        """Add a chart to the worksheet."""
        chart_type = chart_config.get('type', 'bar')
        data_range = chart_config.get('data_range')
        title = chart_config.get('title', 'Chart')
        position = chart_config.get('position', 'G2')
        
        if not data_range:
            return
        
        # Create chart based on type
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            chart = BarChart()  # Default to bar chart
        
        # Set chart properties
        chart.title = title
        chart.style = 10
        
        # Add data to chart
        data = Reference(sheet, range_string=data_range)
        chart.add_data(data, titles_from_data=True)
        
        # Add chart to sheet
        sheet.add_chart(chart, position)
    
    def _auto_adjust_columns(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Auto-adjust column widths based on content."""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_default_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Create a default sheet with sample data."""
        sheet = workbook.create_sheet(title="Data")
        
        # Add sample headers
        headers = ["Item", "Value", "Category"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add sample data
        sample_data = [
            ["Sample Item 1", 100, "Category A"],
            ["Sample Item 2", 200, "Category B"],
            ["Sample Item 3", 150, "Category A"]
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        self._auto_adjust_columns(sheet)
    
    def _save_workbook(self, workbook: openpyxl.Workbook, filename: str) -> str:
        """Save the workbook and return the file path."""
        # Ensure filename has .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Create temp directory if it doesn't exist
        temp_dir = os.path.join(os.getcwd(), 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        
        # Save file
        output_path = os.path.join(temp_dir, filename)
        workbook.save(output_path)
        
        return output_path